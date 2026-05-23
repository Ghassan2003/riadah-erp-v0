"""
Utility functions for the ERP system.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from io import BytesIO


def export_to_excel(data, columns, sheet_name='بيانات', filename='export.xlsx', title=None):
    """
    Enhanced Export data to Excel file with professional formatting.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.rightToLeft = True

    # Styles
    title_font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Arial', size=10)
    cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    total_font = Font(name='Arial', bold=True, size=10, color='1F4E79')
    total_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    even_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    row_start = 1

    # Title row
    if title:
        ws.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=len(columns))
        cell = ws.cell(row=row_start, column=1, value=title)
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for c in range(2, len(columns) + 1):
            ws.cell(row=row_start, column=c).fill = title_fill
        row_start += 1

    # Write headers
    for col_idx, (key, header, width) in enumerate(columns, 1):
        cell = ws.cell(row=row_start, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Write data rows
    numeric_cols = set()
    for row_idx, row_data in enumerate(data):
        excel_row = row_idx + row_start + 1
        for col_idx, (key, header, width) in enumerate(columns, 1):
            value = row_data.get(key, '')
            if value is None: value = ''
            if isinstance(value, bool): value = 'نعم' if value else 'لا'

            # Detect numeric columns
            if isinstance(value, (int, float)):
                numeric_cols.add(col_idx)

            cell = ws.cell(row=excel_row, column=col_idx, value=str(value) if not isinstance(value, (int, float)) else value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Zebra striping
            if row_idx % 2 == 1:
                cell.fill = even_fill

    # Total row for numeric columns
    if data and numeric_cols:
        total_row = len(data) + row_start + 1
        for col_idx, (key, header, width) in enumerate(columns, 1):
            cell = ws.cell(row=total_row, column=col_idx)
            if col_idx in numeric_cols:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                cell.value = f'=SUM({col_letter}{row_start+1}:{col_letter}{total_row-1})'
                cell.number_format = '#,##0.00'
            else:
                cell.value = 'الإجمالي'
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Auto-filter
    last_row = len(data) + row_start + (1 if data and numeric_cols else 0)
    ws.auto_filter.ref = f'A{row_start}:{openpyxl.utils.get_column_letter(len(columns))}{last_row}'

    # Freeze panes (header)
    ws.freeze_panes = f'A{row_start + 1}'

    # Save
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def prepare_export_data(queryset, fields_map):
    """
    Convert queryset to list of dicts for export.

    Args:
        queryset: Django queryset
        fields_map: dict mapping field names to display names

    Returns:
        list of dicts
    """
    data = []
    for obj in queryset:
        row = {}
        for field_key, field_path in fields_map.items():
            value = obj
            for attr in field_path.split('.'):
                if value is None:
                    value = ''
                    break
                if hasattr(value, attr):
                    value = getattr(value, attr)
                else:
                    value = ''
                    break
            # Handle method calls (like get_status_display)
            if callable(value) and not hasattr(value, '__self__'):
                pass
            elif callable(value):
                value = value()
            row[field_key] = value
        data.append(row)
    return data
