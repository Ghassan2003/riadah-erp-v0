from datetime import datetime, timedelta

from django.db.models import Sum, Count, Q, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone

from rest_framework import status, filters
from rest_framework.generics import (
    ListAPIView, CreateAPIView, RetrieveAPIView,
    UpdateAPIView, DestroyAPIView
)
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Invoice, InvoiceItem, Payment, PaymentReminder
from .serializers import (
    InvoiceListSerializer, InvoiceCreateSerializer, InvoiceUpdateSerializer,
    InvoiceDetailSerializer, InvoiceItemSerializer, InvoiceItemCreateSerializer,
    PaymentListSerializer, PaymentCreateSerializer,
    PaymentReminderListSerializer, PaymentReminderCreateSerializer,
    InvoiceStatsSerializer, InvoiceChangeStatusSerializer,
)


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


# ─── Invoice Views ──────────────────────────────────────────────────────────

class InvoiceListView(ListAPIView):
    serializer_class = InvoiceListSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]

    def get_queryset(self):
        queryset = Invoice.objects.filter(is_active=True)

        # Filter by invoice_type
        invoice_type = self.request.query_params.get('invoice_type')
        if invoice_type:
            queryset = queryset.filter(invoice_type=invoice_type)

        # Filter by status
        inv_status = self.request.query_params.get('status')
        if inv_status:
            queryset = queryset.filter(status=inv_status)

        # Filter by payment_status
        payment_status = self.request.query_params.get('payment_status')
        if payment_status:
            queryset = queryset.filter(payment_status=payment_status)

        # Filter by customer
        customer_id = self.request.query_params.get('customer')
        if customer_id:
            queryset = queryset.filter(customer_id=customer_id)

        # Filter by date range
        date_from = self.request.query_params.get('date_from')
        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(issue_date__gte=date_from)
            except ValueError:
                pass

        date_to = self.request.query_params.get('date_to')
        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(issue_date__lte=date_to)
            except ValueError:
                pass

        # Search by invoice_number or customer/supplier name
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(invoice_number__icontains=search) |
                Q(customer__name__icontains=search) |
                Q(supplier__name__icontains=search)
            )

        # Ordering
        ordering = self.request.query_params.get('ordering', '-created_at')
        if ordering in ['invoice_number', '-invoice_number', 'issue_date', '-issue_date',
                        'due_date', '-due_date', 'total_amount', '-total_amount',
                        'created_at', '-created_at']:
            queryset = queryset.order_by(ordering)
        else:
            queryset = queryset.order_by('-created_at')

        return queryset

    search_fields = ['invoice_number', 'customer__name', 'supplier__name']
    ordering_fields = ['invoice_number', 'issue_date', 'due_date', 'total_amount', 'created_at']


class InvoiceCreateView(CreateAPIView):
    serializer_class = InvoiceCreateSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class InvoiceDetailView(RetrieveAPIView):
    serializer_class = InvoiceDetailSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Invoice.objects.filter(is_active=True)


class InvoiceUpdateView(UpdateAPIView):
    serializer_class = InvoiceUpdateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Invoice.objects.filter(is_active=True)


class InvoiceDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            invoice = Invoice.objects.get(pk=pk, is_active=True)
        except Invoice.DoesNotExist:
            return Response(
                {'error': 'الفاتورة غير موجودة'},
                status=status.HTTP_404_NOT_FOUND
            )

        invoice.is_active = False
        invoice.status = 'cancelled'
        invoice.save(update_fields=['is_active', 'status', 'updated_at'])

        return Response(
            {'message': 'تم حذف الفاتورة بنجاح'},
            status=status.HTTP_200_OK
        )


class InvoiceRestoreView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            invoice = Invoice.objects.get(pk=pk, is_active=False)
        except Invoice.DoesNotExist:
            return Response(
                {'error': 'الفاتورة غير موجودة أو ليست محذوفة'},
                status=status.HTTP_404_NOT_FOUND
            )

        invoice.is_active = True
        invoice.status = 'draft'
        invoice.save(update_fields=['is_active', 'status', 'updated_at'])

        return Response(
            {'message': 'تم استعادة الفاتورة بنجاح'},
            status=status.HTTP_200_OK
        )


class InvoiceChangeStatusView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            invoice = Invoice.objects.get(pk=pk, is_active=True)
        except Invoice.DoesNotExist:
            return Response(
                {'error': 'الفاتورة غير موجودة'},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = InvoiceChangeStatusSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        action = serializer.validated_data['action']

        if action == 'send':
            if invoice.status != 'draft':
                return Response(
                    {'error': 'يمكن إرسال الفاتورة فقط من حالة المسودة'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            invoice.status = 'sent'

        elif action == 'accept':
            if invoice.status != 'sent':
                return Response(
                    {'error': 'يمكن قبول الفاتورة فقط من حالة المرسلة'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            invoice.status = 'accepted'

        elif action == 'cancel':
            if invoice.status in ('paid', 'partially_paid'):
                return Response(
                    {'error': 'لا يمكن إلغاء فاتورة مدفوعة أو مدفوعة جزئياً'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            invoice.status = 'cancelled'
            invoice.payment_status = 'cancelled'

        else:
            return Response(
                {'error': 'إجراء غير صالح'},
                status=status.HTTP_400_BAD_REQUEST
            )

        invoice.save(update_fields=['status', 'payment_status', 'updated_at'])

        return Response(
            {
                'message': f'تم تغيير حالة الفاتورة بنجاح',
                'status': invoice.status,
                'payment_status': invoice.payment_status
            },
            status=status.HTTP_200_OK
        )


class InvoiceSendView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            invoice = Invoice.objects.get(pk=pk, is_active=True)
        except Invoice.DoesNotExist:
            return Response(
                {'error': 'الفاتورة غير موجودة'},
                status=status.HTTP_404_NOT_FOUND
            )

        if invoice.status != 'draft':
            return Response(
                {'error': 'يمكن إرسال الفاتورة فقط من حالة المسودة'},
                status=status.HTTP_400_BAD_REQUEST
            )

        invoice.status = 'sent'
        invoice.save(update_fields=['status', 'updated_at'])

        return Response(
            {
                'message': 'تم إرسال الفاتورة بنجاح',
                'status': invoice.status
            },
            status=status.HTTP_200_OK
        )


class InvoiceDuplicateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            original = Invoice.objects.get(pk=pk, is_active=True)
        except Invoice.DoesNotExist:
            return Response(
                {'error': 'الفاتورة غير موجودة'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Create duplicate invoice
        duplicate = Invoice(
            invoice_type=original.invoice_type,
            customer=original.customer,
            supplier=original.supplier,
            issue_date=timezone.now().date(),
            due_date=timezone.now().date() + timedelta(days=30),
            subtotal=original.subtotal,
            discount_type=original.discount_type,
            discount_value=original.discount_value,
            discount_amount=original.discount_amount,
            vat_rate=original.vat_rate,
            vat_amount=original.vat_amount,
            total_after_discount=original.total_after_discount,
            total_amount=original.total_amount,
            currency=original.currency,
            tax_number=original.tax_number,
            company_tax_number=original.company_tax_number,
            notes=original.notes,
            terms_conditions=original.terms_conditions,
            created_by=request.user,
            status='draft',
            payment_status='unpaid',
            paid_amount=0,
        )
        duplicate.save()

        # Copy items
        for item in original.items.all():
            InvoiceItem.objects.create(
                invoice=duplicate,
                product=item.product,
                description=item.description,
                quantity=item.quantity,
                unit_price=item.unit_price,
                unit=item.unit,
                discount_type=item.discount_type,
                discount_value=item.discount_value,
                vat_rate=item.vat_rate,
            )

        # Recalculate to ensure totals are correct
        duplicate.recalculate_totals()

        return Response(
            {
                'message': 'تم نسخ الفاتورة بنجاح',
                'invoice_id': duplicate.id,
                'invoice_number': duplicate.invoice_number
            },
            status=status.HTTP_201_CREATED
        )


# ─── Payment Views ──────────────────────────────────────────────────────────

class PaymentListView(ListAPIView):
    serializer_class = PaymentListSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        pk = self.kwargs.get('pk')
        return Payment.objects.filter(invoice_id=pk).order_by('-created_at')


class PaymentCreateView(CreateAPIView):
    serializer_class = PaymentCreateSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class PaymentDetailView(RetrieveAPIView):
    serializer_class = PaymentListSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Payment.objects.select_related('invoice')


class PaymentUpdateView(UpdateAPIView):
    serializer_class = PaymentCreateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Payment.objects.select_related('invoice')


class PaymentDeleteView(DestroyAPIView):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Payment.objects.all()

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        # The delete() method on Payment model already handles invoice update
        self.perform_destroy(instance)
        return Response(
            {'message': 'تم حذف الدفعة بنجاح'},
            status=status.HTTP_200_OK
        )


# ─── Reminder Views ─────────────────────────────────────────────────────────

class ReminderListView(ListAPIView):
    serializer_class = PaymentReminderListSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        pk = self.kwargs.get('pk')
        return PaymentReminder.objects.filter(invoice_id=pk).order_by('-sent_date')


class ReminderDetailView(RetrieveAPIView):
    serializer_class = PaymentReminderListSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return PaymentReminder.objects.select_related('invoice')


class ReminderUpdateView(UpdateAPIView):
    serializer_class = PaymentReminderCreateSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return PaymentReminder.objects.all()


class ReminderDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            reminder = PaymentReminder.objects.get(pk=pk)
        except PaymentReminder.DoesNotExist:
            return Response(
                {'error': 'تذكير الدفع غير موجود'},
                status=status.HTTP_404_NOT_FOUND
            )

        reminder.delete()

        return Response(
            {'message': 'تم حذف تذكير الدفع بنجاح'},
            status=status.HTTP_200_OK
        )


class ReminderCreateView(CreateAPIView):
    serializer_class = PaymentReminderCreateSerializer
    permission_classes = [IsAuthenticated]


# ─── Stats & Export Views ────────────────────────────────────────────────────

class InvoiceStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        queryset = Invoice.objects.filter(is_active=True)

        # Overall stats
        total_invoices = queryset.count()
        total_sales = queryset.filter(invoice_type='sales').aggregate(
            total=Coalesce(Sum('total_amount'), 0, output_field=DecimalField())
        )['total']
        total_purchases = queryset.filter(invoice_type='purchase').aggregate(
            total=Coalesce(Sum('total_amount'), 0, output_field=DecimalField())
        )['total']
        total_revenue = queryset.filter(invoice_type='sales').aggregate(
            total=Coalesce(Sum('paid_amount'), 0, output_field=DecimalField())
        )['total']
        total_paid = queryset.filter(payment_status='paid').aggregate(
            total=Coalesce(Sum('total_amount'), 0, output_field=DecimalField())
        )['total']
        total_unpaid = queryset.filter(payment_status='unpaid').aggregate(
            total=Coalesce(Sum('total_amount'), 0, output_field=DecimalField())
        )['total']
        total_overdue = queryset.filter(payment_status='overdue').aggregate(
            total=Coalesce(Sum('total_amount'), 0, output_field=DecimalField())
        )['total']

        # Counts by status
        paid_count = queryset.filter(payment_status='paid').count()
        unpaid_count = queryset.filter(payment_status='unpaid').count()
        partially_paid_count = queryset.filter(payment_status='partially_paid').count()
        overdue_count = queryset.filter(payment_status='overdue').count()

        stats = {
            'total_invoices': total_invoices,
            'total_sales': total_sales,
            'total_purchases': total_purchases,
            'total_revenue': total_revenue,
            'total_paid': total_paid,
            'total_unpaid': total_unpaid,
            'total_overdue': total_overdue,
            'paid_count': paid_count,
            'unpaid_count': unpaid_count,
            'partially_paid_count': partially_paid_count,
            'overdue_count': overdue_count,
        }

        serializer = InvoiceStatsSerializer(stats)
        return Response(serializer.data, status=status.HTTP_200_OK)


class InvoiceExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response(
                {'error': 'يرجى تثبيت مكتبة openpyxl: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        queryset = Invoice.objects.filter(is_active=True).order_by('-created_at')

        # Apply filters (same as list view)
        invoice_type = request.query_params.get('invoice_type')
        if invoice_type:
            queryset = queryset.filter(invoice_type=invoice_type)

        inv_status = request.query_params.get('status')
        if inv_status:
            queryset = queryset.filter(status=inv_status)

        payment_status = request.query_params.get('payment_status')
        if payment_status:
            queryset = queryset.filter(payment_status=payment_status)

        date_from = request.query_params.get('date_from')
        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(issue_date__gte=date_from)
            except ValueError:
                pass

        date_to = request.query_params.get('date_to')
        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(issue_date__lte=date_to)
            except ValueError:
                pass

        invoices = queryset.select_related('customer', 'supplier').prefetch_related('items')

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'الفواتير'
        ws.sheet_view.rightToLeft = True

        # Styles
        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            'رقم الفاتورة', 'نوع الفاتورة', 'العميل', 'المورد',
            'تاريخ الإصدار', 'تاريخ الاستحقاق', 'المجموع الفرعي',
            'الخصم', 'الضريبة', 'الإجمالي', 'المدفوع', 'المتبقي',
            'حالة الدفع', 'الحالة', 'العملة'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        INVOICE_TYPE_MAP = dict(Invoice.INVOICE_TYPE_CHOICES)
        PAYMENT_STATUS_MAP = dict(Invoice.PAYMENT_STATUS_CHOICES)
        STATUS_MAP = dict(Invoice.STATUS_CHOICES)

        for row_idx, invoice in enumerate(invoices, 2):
            data = [
                invoice.invoice_number,
                INVOICE_TYPE_MAP.get(invoice.invoice_type, invoice.invoice_type),
                invoice.customer.name if invoice.customer else '',
                invoice.supplier.name if invoice.supplier else '',
                str(invoice.issue_date),
                str(invoice.due_date),
                float(invoice.subtotal),
                float(invoice.discount_amount),
                float(invoice.vat_amount),
                float(invoice.total_amount),
                float(invoice.paid_amount),
                float(invoice.remaining_amount),
                PAYMENT_STATUS_MAP.get(invoice.payment_status, invoice.payment_status),
                STATUS_MAP.get(invoice.status, invoice.status),
                invoice.currency,
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Auto-width columns
        for col in range(1, len(headers) + 1):
            max_length = len(str(headers[col - 1]))
            for row in range(2, len(invoices) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 4, 30)

        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="invoices_export.xlsx"'
        wb.save(response)

        return response
