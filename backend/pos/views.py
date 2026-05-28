from django.utils import timezone
from django.db import transaction as db_transaction
from django.db.models import Sum, Count, Q, F, DecimalField
from django.db.models.functions import TruncHour
from django.http import HttpResponse
from rest_framework import status, generics, pagination
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from collections import Counter

from .models import (
    POSShift,
    POSSale,
    POSRefund,
    POSHoldOrder,
    CashDrawerTransaction,
    PriceList,
    DiscountRule,
    PromoCode,
    LoyaltyProgram,
    LoyaltyTransaction,
    RestaurantTable,
    InstallmentPlan,
    InstallmentPayment,
)
from .serializers import (
    POSShiftListSerializer,
    POSShiftCreateSerializer,
    POSShiftCloseSerializer,
    POSShiftDetailSerializer,
    POSSaleListSerializer,
    POSSaleCreateSerializer,
    POSSaleVoidSerializer,
    POSRefundSerializer,
    POSRefundCreateSerializer,
    POSHoldOrderSerializer,
    POSHoldOrderRecoverSerializer,
    CashDrawerTransactionSerializer,
    CashDrawerTransactionCreateSerializer,
    # New serializers
    PriceListListSerializer,
    PriceListCreateSerializer,
    PriceListUpdateSerializer,
    DiscountRuleListSerializer,
    DiscountRuleCreateSerializer,
    DiscountRuleUpdateSerializer,
    DiscountRuleApplySerializer,
    PromoCodeListSerializer,
    PromoCodeCreateSerializer,
    PromoCodeUpdateSerializer,
    PromoCodeValidateSerializer,
    LoyaltyProgramListSerializer,
    LoyaltyProgramCreateSerializer,
    LoyaltyProgramUpdateSerializer,
    LoyaltyTransactionListSerializer,
    LoyaltyBalanceSerializer,
    LoyaltyRedeemSerializer,
    RestaurantTableListSerializer,
    RestaurantTableCreateSerializer,
    RestaurantTableUpdateSerializer,
    RestaurantTableStatusSerializer,
    InstallmentPlanListSerializer,
    InstallmentPlanCreateSerializer,
    InstallmentPlanUpdateSerializer,
    InstallmentPaymentListSerializer,
    InstallmentPaymentCreateSerializer,
    InstallmentPaymentUpdateSerializer,
)


# ─────────────────────────────────────────────
# Pagination
# ─────────────────────────────────────────────
class StandardPagination(pagination.PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


# ─────────────────────────────────────────────
# Shift Views
# ─────────────────────────────────────────────
class POSShiftListView(generics.ListAPIView):
    """قائمة الورديات"""
    queryset = POSShift.objects.all()
    serializer_class = POSShiftListSerializer
    pagination_class = StandardPagination
    permission_classes = [IsAuthenticated]
    filterset_fields = ['status', 'cashier']
    search_fields = ['shift_number']
    ordering_fields = ['created_at', 'start_time', 'total_sales']
    ordering = ['-created_at']


class POSShiftOpenView(APIView):
    """فتح وردية جديدة"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = POSShiftCreateSerializer(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        shift = serializer.save()

        # Create opening drawer transaction
        CashDrawerTransaction.objects.create(
            shift=shift,
            transaction_type='opening',
            amount=shift.opening_cash,
            description=f'فتح الوردية {shift.shift_number}',
            created_by=request.user,
        )

        detail_serializer = POSShiftDetailSerializer(shift)
        return Response(detail_serializer.data, status=status.HTTP_201_CREATED)


class POSShiftCloseView(APIView):
    """إغلاق الوردية"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            shift = POSShift.objects.get(pk=pk)
        except POSShift.DoesNotExist:
            return Response(
                {'error': 'الوردية غير موجودة'},
                status=status.HTTP_404_NOT_FOUND
            )

        if shift.status == 'closed':
            return Response(
                {'error': 'الوردية مغلقة بالفعل'},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = POSShiftCloseSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        closing_cash = serializer.validated_data['closing_cash']
        notes = serializer.validated_data['notes']

        shift.close_shift(closing_cash, notes)

        # Create closing drawer transaction
        CashDrawerTransaction.objects.create(
            shift=shift,
            transaction_type='closing',
            amount=shift.closing_cash,
            description=f'إغلاق الوردية {shift.shift_number}',
            created_by=request.user,
        )

        detail_serializer = POSShiftDetailSerializer(shift)
        return Response(detail_serializer.data, status=status.HTTP_200_OK)


class POSShiftDetailView(generics.RetrieveAPIView):
    """تفاصيل الوردية"""
    queryset = POSShift.objects.all()
    serializer_class = POSShiftDetailSerializer
    permission_classes = [IsAuthenticated]


# ─────────────────────────────────────────────
# Sale Views
# ─────────────────────────────────────────────
class POSSaleListView(generics.ListAPIView):
    """قائمة عمليات البيع"""
    queryset = POSSale.objects.all()
    serializer_class = POSSaleListSerializer
    pagination_class = StandardPagination
    permission_classes = [IsAuthenticated]
    filterset_fields = ['shift', 'status', 'payment_method', 'customer']
    search_fields = ['receipt_number', 'notes']
    ordering_fields = ['created_at', 'total_amount']
    ordering = ['-created_at']

    def get_queryset(self):
        queryset = super().get_queryset()
        # Filter by date
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')

        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)

        return queryset


class POSSaleCreateView(APIView):
    """إنشاء عملية بيع جديدة"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # Product validation disabled (inventory module not available)

        serializer = POSSaleCreateSerializer(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)

        validated_data = serializer.validated_data
        shift = validated_data['shift']
        items = validated_data['items']

        with db_transaction.atomic():
            # Build item details without inventory validation
            item_details = []
            for item in items:
                # Compute item totals
                line_total = item['quantity'] * item['unit_price']
                item_discount = item.get('discount', 0)
                item_vat = round((line_total - item_discount) * 0.15, 2)
                item_total = line_total - item_discount + item_vat

                item_details.append({
                    'product_id': str(item.get('product_id', '')),
                    'name': item['name'],
                    'quantity': item['quantity'],
                    'unit_price': item['unit_price'],
                    'discount': item_discount,
                    'vat': item_vat,
                    'total': item_total,
                })

            # Create the sale
            sale = POSSale.objects.create(
                shift=shift,
                customer=validated_data.get('customer'),
                items=[
                    {
                        'product_id': d['product_id'],
                        'name': d['name'],
                        'quantity': d['quantity'],
                        'unit_price': d['unit_price'],
                        'discount': d['discount'],
                        'vat': d['vat'],
                        'total': d['total'],
                    }
                    for d in item_details
                ],
                subtotal=validated_data['subtotal'],
                discount_amount=validated_data['discount_amount'],
                vat_amount=validated_data['vat_amount'],
                total_amount=validated_data['total_amount'],
                payment_method=validated_data.get('payment_method', 'cash'),
                cash_received=validated_data.get('cash_received', 0),
                change_amount=validated_data.get('change_amount', 0),
                card_last_four=validated_data.get('card_last_four', ''),
                notes=validated_data.get('notes', ''),
            )

            # Stock deduction not available (inventory module removed)

            # Update shift totals
            shift.total_sales = (shift.total_sales or 0) + sale.total_amount
            shift.total_transactions = (shift.total_transactions or 0) + 1
            shift.save(update_fields=['total_sales', 'total_transactions'])

        return Response({
            'id': sale.id,
            'receipt_number': sale.receipt_number,
            'total_amount': str(sale.total_amount),
            'change_amount': str(sale.change_amount),
            'payment_method': sale.payment_method,
            'status': sale.status,
            'created_at': sale.created_at,
        }, status=status.HTTP_201_CREATED)


class POSSaleDetailView(generics.RetrieveAPIView):
    """تفاصيل عملية البيع"""
    queryset = POSSale.objects.all()
    serializer_class = POSSaleListSerializer
    permission_classes = [IsAuthenticated]


class POSSaleVoidView(APIView):
    """إلغاء عملية البيع"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            sale = POSSale.objects.get(pk=pk)
        except POSSale.DoesNotExist:
            return Response(
                {'error': 'عملية البيع غير موجودة'},
                status=status.HTTP_404_NOT_FOUND
            )

        if sale.status == 'voided':
            return Response(
                {'error': 'عملية البيع ملغاة بالفعل'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if sale.status == 'refunded':
            return Response(
                {'error': 'تم إرجاع عملية البيع بالفعل، لا يمكن إلغاؤها'},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = POSSaleVoidSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        void_reason = serializer.validated_data['void_reason']

        # Stock restoration not available (inventory module removed)

        # Void the sale
        sale.void_sale(request.user, void_reason)

        # Update shift totals
        shift = sale.shift
        shift.total_sales = max((shift.total_sales or 0) - sale.total_amount, 0)
        shift.total_transactions = max((shift.total_transactions or 0) - 1, 0)
        shift.save(update_fields=['total_sales', 'total_transactions'])

        return Response({
            'id': sale.id,
            'receipt_number': sale.receipt_number,
            'status': sale.status,
            'void_reason': sale.void_reason,
            'voided_by': request.user.get_full_name() or request.user.username,
            'voided_at': sale.voided_at,
        }, status=status.HTTP_200_OK)


# ─────────────────────────────────────────────
# Refund Views
# ─────────────────────────────────────────────
class POSRefundListView(generics.ListAPIView):
    """قائمة الإرجاعات"""
    queryset = POSRefund.objects.all()
    serializer_class = POSRefundSerializer
    pagination_class = StandardPagination
    permission_classes = [IsAuthenticated]
    filterset_fields = ['sale', 'shift', 'refund_method', 'processed_by']
    search_fields = ['refund_number', 'reason']
    ordering_fields = ['created_at', 'refund_amount']
    ordering = ['-created_at']


class POSRefundCreateView(APIView):
    """إنشاء إرجاع جديد"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # Stock restoration not available (inventory module removed)

        serializer = POSRefundCreateSerializer(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)

        validated_data = serializer.validated_data
        sale = validated_data['sale']
        shift = validated_data['shift']
        items = validated_data['items']
        refund_method = validated_data.get('refund_method', 'original')

        # If original, use the sale's payment method
        if refund_method == 'original':
            refund_method = sale.payment_method

        # Stock restoration not available (inventory module removed)

        # Create refund
        refund = POSRefund.objects.create(
            sale=sale,
            shift=shift,
            items=items,
            refund_amount=validated_data['refund_amount'],
            refund_method=refund_method,
            reason=validated_data['reason'],
            processed_by=validated_data.get('processed_by') or request.user,
        )

        # Mark sale as refunded
        sale.status = 'refunded'
        sale.save(update_fields=['status'])

        # Update shift totals
        shift.total_refunds = (shift.total_refunds or 0) + refund.refund_amount
        shift.save(update_fields=['total_refunds'])

        refund_serializer = POSRefundSerializer(refund)
        return Response(refund_serializer.data, status=status.HTTP_201_CREATED)


class POSRefundDetailView(generics.RetrieveAPIView):
    """تفاصيل الإرجاع"""
    queryset = POSRefund.objects.all()
    serializer_class = POSRefundSerializer
    permission_classes = [IsAuthenticated]


# ─────────────────────────────────────────────
# Hold Order Views
# ─────────────────────────────────────────────
class POSHoldListView(generics.ListAPIView):
    """قائمة الطلبات المعلقة"""
    queryset = POSHoldOrder.objects.filter(is_recovered=False)
    serializer_class = POSHoldOrderSerializer
    pagination_class = StandardPagination
    permission_classes = [IsAuthenticated]
    filterset_fields = ['shift']
    ordering = ['-created_at']


class POSHoldCreateView(generics.CreateAPIView):
    """إنشاء طلب معلق"""
    queryset = POSHoldOrder.objects.all()
    serializer_class = POSHoldOrderSerializer
    permission_classes = [IsAuthenticated]


class POSHoldDeleteView(generics.DestroyAPIView):
    """حذف طلب معلق"""
    queryset = POSHoldOrder.objects.all()
    permission_classes = [IsAuthenticated]


class POSHoldDetailView(generics.RetrieveAPIView):
    """تفاصيل الطلب المعلق"""
    queryset = POSHoldOrder.objects.all()
    serializer_class = POSHoldOrderSerializer
    permission_classes = [IsAuthenticated]


# ─────────────────────────────────────────────
# Cash Drawer Views
# ─────────────────────────────────────────────
class CashDrawerTransactionListView(generics.ListAPIView):
    """قائمة حركات الصندوق"""
    queryset = CashDrawerTransaction.objects.all()
    serializer_class = CashDrawerTransactionSerializer
    pagination_class = StandardPagination
    permission_classes = [IsAuthenticated]
    filterset_fields = ['shift', 'transaction_type', 'created_by']
    ordering_fields = ['created_at']
    ordering = ['-created_at']


class CashDrawerTransactionCreateView(generics.CreateAPIView):
    """إنشاء حركة صندوق (إيداع/سحب)"""
    queryset = CashDrawerTransaction.objects.all()
    serializer_class = CashDrawerTransactionCreateSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class CashDrawerTransactionDetailView(generics.RetrieveAPIView):
    """تفاصيل حركة الصندوق"""
    queryset = CashDrawerTransaction.objects.all()
    serializer_class = CashDrawerTransactionSerializer
    permission_classes = [IsAuthenticated]


# ─────────────────────────────────────────────
# Stats & Export Views
# ─────────────────────────────────────────────
class POSStatsView(APIView):
    """إحصائيات نقطة البيع"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = timezone.now().date()
        today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)

        # Today's sales
        today_sales = POSSale.objects.filter(
            created_at__date=today,
            status='completed',
        )
        today_sales_total = today_sales.aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        today_sales_count = today_sales.count()

        # Today's refunds
        today_refunds = POSRefund.objects.filter(created_at__date=today)
        today_refunds_total = today_refunds.aggregate(
            total=Sum('refund_amount')
        )['total'] or 0

        today_net_sales = today_sales_total - today_refunds_total

        # Payment method breakdown
        payment_breakdown = today_sales.values('payment_method').annotate(
            total=Sum('total_amount'),
            count=Count('id'),
        ).order_by('-total')

        payment_method_breakdown = {}
        for item in payment_breakdown:
            payment_method_breakdown[item['payment_method']] = {
                'total': str(item['total']),
                'count': item['count'],
            }

        # Top products today
        all_items = []
        for sale in today_sales:
            if sale.items:
                for item in sale.items:
                    all_items.append({
                        'product_id': item.get('product_id'),
                        'name': item.get('name', ''),
                        'quantity': item.get('quantity', 0),
                        'total': item.get('total', 0),
                    })

        # Aggregate by product
        product_totals = {}
        for item in all_items:
            pid = item['product_id']
            if pid not in product_totals:
                product_totals[pid] = {
                    'product_id': pid,
                    'name': item['name'],
                    'quantity': 0,
                    'total': 0,
                }
            product_totals[pid]['quantity'] += item['quantity']
            product_totals[pid]['total'] += item['total']

        top_products = sorted(
            product_totals.values(),
            key=lambda x: x['total'],
            reverse=True,
        )[:10]

        # Hourly sales
        hourly_data = today_sales.annotate(
            hour=TruncHour('created_at')
        ).values('hour').annotate(
            total=Sum('total_amount'),
            count=Count('id'),
        ).order_by('hour')

        hourly_sales = []
        for h in hourly_data:
            hourly_sales.append({
                'hour': h['hour'].strftime('%H:00') if h['hour'] else '',
                'total': str(h['total']),
                'count': h['count'],
            })

        # Current open shift for current user
        current_shift = POSShift.objects.filter(
            cashier=request.user,
            status='open',
        ).first()

        current_shift_data = None
        if current_shift:
            current_shift_data = POSShiftDetailSerializer(current_shift).data

        return Response({
            'date': today.isoformat(),
            'today_sales_total': str(today_sales_total),
            'today_sales_count': today_sales_count,
            'today_refunds_total': str(today_refunds_total),
            'today_net_sales': str(today_net_sales),
            'current_shift': current_shift_data,
            'payment_method_breakdown': payment_method_breakdown,
            'top_products': top_products,
            'hourly_sales': hourly_sales,
        })


class POSExportView(APIView):
    """تصدير بيانات نقطة البيع"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            return Response(
                {'error': 'يجب تثبيت مكتبة openpyxl: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # Get filter parameters
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        export_type = request.query_params.get('type', 'sales')

        wb = openpyxl.Workbook()

        # Styles
        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        if export_type == 'sales' or export_type == 'all':
            # Sales sheet
            ws = wb.active
            ws.title = 'عمليات البيع'
            ws.sheet_view.rightToLeft = True

            headers = [
                'رقم الإيصال', 'الكاشير', 'العميل', 'المجموع الفرعي',
                'الخصم', 'ضريبة القيمة المضافة', 'الإجمالي',
                'طريقة الدفع', 'الحالة', 'التاريخ',
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            sales = POSSale.objects.all()
            if date_from:
                sales = sales.filter(created_at__date__gte=date_from)
            if date_to:
                sales = sales.filter(created_at__date__lte=date_to)

            for row, sale in enumerate(sales, 2):
                cashier_name = (
                    sale.shift.cashier.get_full_name() or sale.shift.cashier.username
                )
                customer_name = ''
                if sale.customer:
                    customer_name = getattr(sale.customer, 'name', str(sale.customer))

                data = [
                    sale.receipt_number,
                    cashier_name,
                    customer_name,
                    str(sale.subtotal),
                    str(sale.discount_amount),
                    str(sale.vat_amount),
                    str(sale.total_amount),
                    sale.get_payment_method_display(),
                    sale.get_status_display(),
                    sale.created_at.strftime('%Y-%m-%d %H:%M'),
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border

            # Auto-width
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = max(max_length + 4, 15)

        if export_type == 'shifts' or export_type == 'all':
            # Shifts sheet
            if export_type != 'all':
                ws = wb.active
            else:
                ws = wb.create_sheet('الورديات')

            ws.title = 'الورديات'
            ws.sheet_view.rightToLeft = True

            headers = [
                'رقم الوردية', 'الكاشير', 'وقت البدء', 'وقت الانتهاء',
                'الحالة', 'المبلغ الافتتاحي', 'المبلغ النهائي',
                'المبلغ المتوقع', 'الفرق', 'إجمالي المبيعات',
                'عدد العمليات', 'إجمالي المردودات',
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            shifts = POSShift.objects.all()
            if date_from:
                shifts = shifts.filter(created_at__date__gte=date_from)
            if date_to:
                shifts = shifts.filter(created_at__date__lte=date_to)

            for row, shift_obj in enumerate(shifts, 2):
                cashier_name = shift_obj.cashier.get_full_name() or shift_obj.cashier.username
                data = [
                    shift_obj.shift_number,
                    cashier_name,
                    shift_obj.start_time.strftime('%Y-%m-%d %H:%M') if shift_obj.start_time else '',
                    shift_obj.end_time.strftime('%Y-%m-%d %H:%M') if shift_obj.end_time else '',
                    shift_obj.get_status_display(),
                    str(shift_obj.opening_cash),
                    str(shift_obj.closing_cash),
                    str(shift_obj.expected_cash),
                    str(shift_obj.difference),
                    str(shift_obj.total_sales),
                    str(shift_obj.total_transactions),
                    str(shift_obj.total_refunds),
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = max(max_length + 4, 15)

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename=pos_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        wb.save(response)
        return response


# ============================================================
# Price List Views
# ============================================================


class PriceListListView(generics.ListCreateAPIView):
    """قائمة قوائم الأسعار"""
    queryset = PriceList.objects.all()
    pagination_class = StandardPagination
    permission_classes = [IsAuthenticated]
    filterset_fields = ['is_default', 'is_active']
    search_fields = ['name', 'description']
    ordering_fields = ['created_at', 'name']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return PriceListCreateSerializer
        return PriceListListSerializer


class PriceListDetailView(generics.RetrieveUpdateAPIView):
    """تفاصيل قائمة الأسعار"""
    queryset = PriceList.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return PriceListUpdateSerializer
        return PriceListListSerializer


class PriceListDeleteView(APIView):
    """حذف قائمة أسعار"""
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            price_list = PriceList.objects.get(pk=pk)
        except PriceList.DoesNotExist:
            return Response(
                {'error': 'قائمة الأسعار غير موجودة'},
                status=status.HTTP_404_NOT_FOUND
            )
        price_list.delete()
        return Response(
            {'message': 'تم حذف قائمة الأسعار بنجاح'},
            status=status.HTTP_200_OK
        )


# ============================================================
# Discount Rule Views
# ============================================================


class DiscountRuleListView(generics.ListCreateAPIView):
    """قائمة قواعد الخصم"""
    queryset = DiscountRule.objects.all()
    pagination_class = StandardPagination
    permission_classes = [IsAuthenticated]
    filterset_fields = ['discount_type', 'applies_to', 'is_active']
    search_fields = ['name', 'product_category']
    ordering_fields = ['created_at', 'priority', 'name']
    ordering = ['-priority', '-created_at']

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return DiscountRuleCreateSerializer
        return DiscountRuleListSerializer


class DiscountRuleDetailView(generics.RetrieveUpdateAPIView):
    """تفاصيل قاعدة الخصم"""
    queryset = DiscountRule.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return DiscountRuleUpdateSerializer
        return DiscountRuleListSerializer


class DiscountRuleDeleteView(APIView):
    """حذف قاعدة خصم"""
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            rule = DiscountRule.objects.get(pk=pk)
        except DiscountRule.DoesNotExist:
            return Response(
                {'error': 'قاعدة الخصم غير موجودة'},
                status=status.HTTP_404_NOT_FOUND
            )
        rule.delete()
        return Response(
            {'message': 'تم حذف قاعدة الخصم بنجاح'},
            status=status.HTTP_200_OK
        )


class DiscountRuleApplyView(APIView):
    """البحث عن قواعد الخصم المطبقة على سلة المشتريات"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = DiscountRuleApplySerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)

        product_category = serializer.validated_data.get('product_category', '')
        quantity = serializer.validated_data.get('quantity', 1)
        amount = serializer.validated_data.get('amount', 0)
        today = timezone.now().date()

        # Find applicable active rules
        rules = DiscountRule.objects.filter(is_active=True).filter(
            Q(valid_from__isnull=True) | Q(valid_from__lte=today)
        ).filter(
            Q(valid_until__isnull=True) | Q(valid_until__gte=today)
        ).filter(
            min_quantity__lte=quantity
        ).order_by('-priority')

        applicable_rules = []
        for rule in rules:
            # Check if rule applies to this context
            if rule.applies_to == 'all_products':
                applicable_rules.append(rule)
            elif rule.applies_to == 'specific_category':
                if product_category and product_category == rule.product_category:
                    applicable_rules.append(rule)

        # Calculate discount for each applicable rule
        results = []
        for rule in applicable_rules:
            discount_amount = 0
            if rule.discount_type == 'percentage':
                discount_amount = float(amount) * float(rule.discount_value) / 100
            elif rule.discount_type == 'fixed_amount':
                discount_amount = float(rule.discount_value)
            elif rule.discount_type == 'buy_x_get_y':
                # Simplified: give discount for each qualifying set
                sets = int(quantity) // rule.min_quantity if rule.min_quantity > 0 else 0
                discount_amount = sets * float(rule.discount_value)

            # Cap at max_discount_amount if set
            if rule.max_discount_amount and discount_amount > float(rule.max_discount_amount):
                discount_amount = float(rule.max_discount_amount)

            results.append({
                'id': rule.id,
                'name': rule.name,
                'discount_type': rule.discount_type,
                'discount_type_display': rule.get_discount_type_display(),
                'discount_value': str(rule.discount_value),
                'calculated_discount': str(round(discount_amount, 2)),
                'priority': rule.priority,
            })

        return Response({
            'product_category': product_category,
            'quantity': quantity,
            'amount': str(amount),
            'applicable_rules': results,
            'total_discount': str(round(sum(float(r['calculated_discount']) for r in results), 2)),
        })


# ============================================================
# Promo Code Views
# ============================================================


class PromoCodeListView(generics.ListCreateAPIView):
    """قائمة الأكواد الترويجية"""
    queryset = PromoCode.objects.all()
    pagination_class = StandardPagination
    permission_classes = [IsAuthenticated]
    filterset_fields = ['discount_type', 'is_active']
    search_fields = ['code', 'description']
    ordering_fields = ['created_at', 'code']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return PromoCodeCreateSerializer
        return PromoCodeListSerializer


class PromoCodeDetailView(generics.RetrieveUpdateAPIView):
    """تفاصيل الكود الترويجي"""
    queryset = PromoCode.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return PromoCodeUpdateSerializer
        return PromoCodeListSerializer


class PromoCodeDeleteView(APIView):
    """حذف كود ترويجي"""
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            promo = PromoCode.objects.get(pk=pk)
        except PromoCode.DoesNotExist:
            return Response(
                {'error': 'الكود الترويجي غير موجود'},
                status=status.HTTP_404_NOT_FOUND
            )
        promo.delete()
        return Response(
            {'message': 'تم حذف الكود الترويجي بنجاح'},
            status=status.HTTP_200_OK
        )


class PromoCodeValidateView(APIView):
    """التحقق من صلاحية الكود الترويجي"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = PromoCodeValidateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        code = serializer.validated_data['code'].upper().strip()
        order_amount = float(serializer.validated_data['order_amount'])
        customer_id = serializer.validated_data.get('customer_id')

        try:
            promo = PromoCode.objects.get(code=code)
        except PromoCode.DoesNotExist:
            return Response({
                'is_valid': False,
                'message': 'الكود الترويجي غير موجود',
            }, status=status.HTTP_200_OK)

        # Check active
        if not promo.is_active:
            return Response({
                'is_valid': False,
                'message': 'الكود الترويجي غير مفعّل',
            }, status=status.HTTP_200_OK)

        now = timezone.now()

        # Check dates
        if promo.valid_from and now < promo.valid_from:
            return Response({
                'is_valid': False,
                'message': 'الكود الترويجي لم يبدأ بعد',
            }, status=status.HTTP_200_OK)

        if promo.valid_until and now > promo.valid_until:
            return Response({
                'is_valid': False,
                'message': 'الكود الترويجي منتهي الصلاحية',
            }, status=status.HTTP_200_OK)

        # Check max uses
        if promo.max_uses and promo.used_count >= promo.max_uses:
            return Response({
                'is_valid': False,
                'message': 'تم استخدام الكود الترويجي بالحد الأقصى',
            }, status=status.HTTP_200_OK)

        # Check min order amount
        if promo.min_order_amount and order_amount < float(promo.min_order_amount):
            return Response({
                'is_valid': False,
                'message': f'الحد الأدنى للطلب هو {promo.min_order_amount}',
            }, status=status.HTTP_200_OK)

        # Calculate discount
        discount_amount = 0
        if promo.discount_type == 'percentage':
            discount_amount = order_amount * float(promo.discount_value) / 100
        elif promo.discount_type == 'fixed_amount':
            discount_amount = float(promo.discount_value)
        elif promo.discount_type == 'free_shipping':
            discount_amount = 0

        return Response({
            'is_valid': True,
            'message': 'الكود الترويجي صالح',
            'code': promo.code,
            'discount_type': promo.discount_type,
            'discount_type_display': promo.get_discount_type_display(),
            'discount_value': str(promo.discount_value),
            'calculated_discount': str(round(discount_amount, 2)),
            'final_amount': str(round(order_amount - discount_amount, 2)),
        }, status=status.HTTP_200_OK)


# ============================================================
# Loyalty Program Views
# ============================================================


class LoyaltyProgramListView(generics.ListCreateAPIView):
    """قائمة برامج الولاء"""
    queryset = LoyaltyProgram.objects.all()
    pagination_class = StandardPagination
    permission_classes = [IsAuthenticated]
    filterset_fields = ['is_active']
    search_fields = ['name', 'description']
    ordering_fields = ['created_at', 'name']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return LoyaltyProgramCreateSerializer
        return LoyaltyProgramListSerializer


class LoyaltyProgramDetailView(generics.RetrieveUpdateAPIView):
    """تفاصيل برنامج الولاء"""
    queryset = LoyaltyProgram.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return LoyaltyProgramUpdateSerializer
        return LoyaltyProgramListSerializer


class LoyaltyProgramDeleteView(APIView):
    """حذف برنامج ولاء"""
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            program = LoyaltyProgram.objects.get(pk=pk)
        except LoyaltyProgram.DoesNotExist:
            return Response(
                {'error': 'برنامج الولاء غير موجود'},
                status=status.HTTP_404_NOT_FOUND
            )
        program.delete()
        return Response(
            {'message': 'تم حذف برنامج الولاء بنجاح'},
            status=status.HTTP_200_OK
        )


class LoyaltyTransactionListView(generics.ListAPIView):
    """قائمة معاملات الولاء"""
    queryset = LoyaltyTransaction.objects.all()
    serializer_class = LoyaltyTransactionListSerializer
    pagination_class = StandardPagination
    permission_classes = [IsAuthenticated]
    filterset_fields = ['loyalty_program', 'customer', 'transaction_type']
    search_fields = ['description']
    ordering_fields = ['created_at']
    ordering = ['-created_at']


class LoyaltyBalanceView(APIView):
    """رصيد الولاء للعميل"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        customer_id = request.query_params.get('customer_id')
        if not customer_id:
            return Response(
                {'error': 'يجب تحديد معرّف العميل'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get balance per program
        programs = LoyaltyProgram.objects.filter(is_active=True)
        balances = []
        total_points = 0

        for program in programs:
            # Get latest transaction balance for this customer+program
            latest_tx = LoyaltyTransaction.objects.filter(
                loyalty_program=program,
                customer_id=customer_id,
            ).order_by('-created_at').first()

            balance = latest_tx.balance_after if latest_tx else 0
            total_points += balance

            # Count transactions
            tx_count = LoyaltyTransaction.objects.filter(
                loyalty_program=program,
                customer_id=customer_id,
            ).count()

            earned = LoyaltyTransaction.objects.filter(
                loyalty_program=program,
                customer_id=customer_id,
                transaction_type='earned',
            ).aggregate(total=Sum('points'))['total'] or 0

            redeemed = LoyaltyTransaction.objects.filter(
                loyalty_program=program,
                customer_id=customer_id,
                transaction_type='redeemed',
            ).aggregate(total=Sum('points'))['total'] or 0

            balances.append({
                'program_id': program.id,
                'program_name': program.name,
                'balance': balance,
                'total_earned': earned,
                'total_redeemed': abs(redeemed),
                'points_value': str(program.points_value),
                'redeemable_amount': str(round(balance * float(program.points_value), 2)),
                'transaction_count': tx_count,
            })

        return Response({
            'customer_id': int(customer_id),
            'total_points': total_points,
            'programs': balances,
        })


class LoyaltyRedeemView(APIView):
    """استبدال نقاط الولاء"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = LoyaltyRedeemSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        customer_id = serializer.validated_data['customer_id']
        program_id = serializer.validated_data['program_id']
        points = serializer.validated_data['points']
        sale_id = serializer.validated_data.get('sale_id')

        try:
            program = LoyaltyProgram.objects.get(pk=program_id, is_active=True)
        except LoyaltyProgram.DoesNotExist:
            return Response(
                {'error': 'برنامج الولاء غير موجود أو غير مفعّل'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check minimum redemption
        if points < program.min_redemption_points:
            return Response(
                {'error': f'الحد الأدنى لاستبدال النقاط هو {program.min_redemption_points}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get current balance
        latest_tx = LoyaltyTransaction.objects.filter(
            loyalty_program=program,
            customer_id=customer_id,
        ).order_by('-created_at').first()

        current_balance = latest_tx.balance_after if latest_tx else 0

        if points > current_balance:
            return Response(
                {'error': f'رصيد النقاط غير كافٍ. الرصيد الحالي: {current_balance}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        with db_transaction.atomic():
            new_balance = current_balance - points
            redeemed_amount = points * float(program.points_value)

            LoyaltyTransaction.objects.create(
                loyalty_program=program,
                customer_id=customer_id,
                pos_sale_id=sale_id,
                points=-points,
                balance_after=new_balance,
                transaction_type='redeemed',
                description=f'استبدال {points} نقطة بمبلغ {redeemed_amount}',
            )

        return Response({
            'message': 'تم استبدال النقاط بنجاح',
            'points_redeemed': points,
            'redeemed_amount': str(round(redeemed_amount, 2)),
            'remaining_balance': new_balance,
        }, status=status.HTTP_200_OK)


# ============================================================
# Restaurant Table Views
# ============================================================


class RestaurantTableListView(generics.ListCreateAPIView):
    """قائمة الطاولات"""
    queryset = RestaurantTable.objects.all()
    pagination_class = StandardPagination
    permission_classes = [IsAuthenticated]
    filterset_fields = ['status', 'area']
    search_fields = ['table_number', 'area']
    ordering_fields = ['created_at', 'table_number']
    ordering = ['table_number']

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return RestaurantTableCreateSerializer
        return RestaurantTableListSerializer


class RestaurantTableDetailView(generics.RetrieveUpdateAPIView):
    """تفاصيل الطاولة"""
    queryset = RestaurantTable.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return RestaurantTableUpdateSerializer
        return RestaurantTableListSerializer


class RestaurantTableStatusView(APIView):
    """تحديث حالة الطاولة"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            table = RestaurantTable.objects.get(pk=pk)
        except RestaurantTable.DoesNotExist:
            return Response(
                {'error': 'الطاولة غير موجودة'},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = RestaurantTableStatusSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        new_status = serializer.validated_data['status']
        old_status = table.status
        table.status = new_status
        table.save(update_fields=['status', 'updated_at'])

        return Response({
            'message': 'تم تحديث حالة الطاولة بنجاح',
            'table_id': table.id,
            'table_number': table.table_number,
            'old_status': old_status,
            'old_status_display': table.get_status_display(),
            'new_status': new_status,
            'new_status_display': dict(RestaurantTable.TABLE_STATUS_CHOICES).get(new_status, new_status),
        }, status=status.HTTP_200_OK)


# ============================================================
# Installment Plan Views
# ============================================================


class InstallmentPlanListView(generics.ListCreateAPIView):
    """قائمة خطط الأقساط"""
    queryset = InstallmentPlan.objects.all()
    pagination_class = StandardPagination
    permission_classes = [IsAuthenticated]
    filterset_fields = ['status', 'customer', 'pos_sale']
    search_fields = ['notes']
    ordering_fields = ['created_at', 'next_due_date', 'total_amount']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return InstallmentPlanCreateSerializer
        return InstallmentPlanListSerializer


class InstallmentPlanDetailView(generics.RetrieveUpdateAPIView):
    """تفاصيل خطة الأقساط"""
    queryset = InstallmentPlan.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return InstallmentPlanUpdateSerializer
        return InstallmentPlanListSerializer


class InstallmentStatsView(APIView):
    """إحصائيات خطط الأقساط النشطة"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        plans = InstallmentPlan.objects.filter(status='active')

        total_active = plans.count()
        total_amount = plans.aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        total_paid = plans.aggregate(
            total=Sum('paid_amount')
        )['total'] or 0
        total_remaining = plans.aggregate(
            total=Sum('remaining_amount')
        )['total'] or 0

        # Overdue installments
        today = timezone.now().date()
        overdue_payments = InstallmentPayment.objects.filter(
            installment_plan__status='active',
            status='pending',
            due_date__lt=today,
        )
        overdue_count = overdue_payments.count()
        overdue_amount = overdue_payments.aggregate(
            total=Sum('amount')
        )['total'] or 0

        # By status
        by_status = list(
            InstallmentPlan.objects.values('status').annotate(
                count=Count('id'),
                total=Sum('total_amount'),
                remaining=Sum('remaining_amount'),
            ).order_by('status')
        )

        return Response({
            'active_plans_count': total_active,
            'total_active_amount': str(total_amount),
            'total_paid_amount': str(total_paid),
            'total_remaining_amount': str(total_remaining),
            'overdue_payments_count': overdue_count,
            'overdue_payments_amount': str(overdue_amount),
            'by_status': by_status,
        })


class InstallmentPaymentListView(generics.ListCreateAPIView):
    """قائمة دفعات الأقساط"""
    queryset = InstallmentPayment.objects.all()
    pagination_class = StandardPagination
    permission_classes = [IsAuthenticated]
    filterset_fields = ['installment_plan', 'status', 'payment_method']
    ordering_fields = ['created_at', 'due_date', 'installment_number']
    ordering = ['installment_number']

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return InstallmentPaymentCreateSerializer
        return InstallmentPaymentListSerializer


class InstallmentPaymentDetailView(generics.RetrieveUpdateAPIView):
    """تفاصيل دفعة القسط"""
    queryset = InstallmentPayment.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return InstallmentPaymentUpdateSerializer
        return InstallmentPaymentListSerializer
